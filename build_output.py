"""
build_output.py — Generate the final XLSX with 4 tabs.

All content is driven by the insights dict produced by synthesize_insights.py.
Nothing is hardcoded — this module renders data, it does not define it.

Tabs:
  1. Vendor Analysis      — full classified vendor list
  2. Top 3 Opportunities  — from insights["opportunities"]
  3. Methodology          — process documentation + QA stats
  4. Recommendations      — executive memo from insights["executive_memo"]
"""
from __future__ import annotations

import textwrap

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from config import OUTPUT_XLSX

# ── Color palette ─────────────────────────────────────────────────────────────
HEADER_FILL   = PatternFill("solid", fgColor="1F3864")
HEADER_FONT   = Font(bold=True, color="FFFFFF", size=11)
TERM_FILL     = PatternFill("solid", fgColor="FFCCCC")
CONS_FILL     = PatternFill("solid", fgColor="FFF2CC")
OPT_FILL      = PatternFill("solid", fgColor="E2EFDA")
WARN_FILL     = PatternFill("solid", fgColor="FCE4D6")
ALT_FILL      = PatternFill("solid", fgColor="F5F5F5")
SECTION_FONT  = Font(bold=True, size=12, color="1F3864")
TITLE_FONT    = Font(bold=True, size=14, color="1F3864")
BODY_FONT     = Font(size=11)
BOLD_FONT     = Font(bold=True, size=11)
MEMO_HDR_FONT = Font(bold=True, size=11)

REC_FILLS = {"Terminate": TERM_FILL, "Consolidate": CONS_FILL, "Optimize": OPT_FILL}


def _set_header_row(ws, headers: list[str], widths: list[int]) -> None:
    ws.append(headers)
    for col_idx, (cell, width) in enumerate(zip(ws[1], widths), 1):
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[1].height = 28


# ── Tab 1: Vendor Analysis ────────────────────────────────────────────────────
def _build_vendor_tab(wb: openpyxl.Workbook, vendors: list[dict], classifications: list[dict]) -> None:
    ws = wb.active
    ws.title = "Vendor Analysis"

    headers = ["Vendor Name", "Department", "Last 12mo Cost (USD)",
               "Description", "Recommendation", "Recommendation Note", "QA Flag"]
    _set_header_row(ws, headers, [42, 22, 22, 52, 16, 48, 14])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = "A1:G1"

    class_map = {c["vendor_name"]: c for c in classifications}
    for row_num, v in enumerate(sorted(vendors, key=lambda x: x["cost_usd"], reverse=True), 2):
        name = v["vendor_name"]
        c = class_map.get(name, {})
        rec = c.get("recommendation", "Optimize")

        qa_flag = ""
        if c.get("qa_reclassified"):
            qa_flag = "QA Revised"
        elif c.get("qa_warn"):
            qa_flag = "QA Warning"

        ws.append([name, c.get("department", ""), v["cost_usd"],
                   c.get("description", ""), rec,
                   c.get("recommendation_note", ""), qa_flag])

        for col_idx, cell in enumerate(ws[row_num], 1):
            cell.font = BODY_FONT
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if qa_flag == "QA Warning" and col_idx == 7:
                cell.fill = WARN_FILL
            elif REC_FILLS.get(rec):
                cell.fill = REC_FILLS[rec]
            elif row_num % 2 == 0:
                cell.fill = ALT_FILL

        ws[f"C{row_num}"].number_format = '"$"#,##0'
        ws.row_dimensions[row_num].height = 40


# ── Tab 2: Top 3 Opportunities ────────────────────────────────────────────────
def _build_opportunities_tab(wb: openpyxl.Workbook, insights: dict) -> None:
    ws = wb.create_sheet("Top 3 Opportunities")
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 44
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 26
    ws.column_dimensions["F"].width = 18

    ws.append(["", "TOP 3 COST REDUCTION OPPORTUNITIES"])
    ws["B1"].font = TITLE_FONT
    ws.row_dimensions[1].height = 30
    ws.append([""])

    headers = ["#", "Opportunity", "Description & Actions", "Current Spend", "Est. Annual Savings", "Timeline"]
    ws.append(headers)
    for cell in ws[3]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[3].height = 26

    opp_fills = [OPT_FILL, OPT_FILL, CONS_FILL]
    opportunities = insights.get("opportunities", [])

    for i, opp in enumerate(opportunities[:3]):
        savings_low  = opp.get("savings_low_usd", 0)
        savings_high = opp.get("savings_high_usd", 0)
        current      = opp.get("current_spend_usd", 0)
        vendors_list = ", ".join(opp.get("affected_vendors", [])[:4])

        desc_body = opp.get("description", "")
        steps     = opp.get("implementation_steps", "")
        risks     = opp.get("risks", "")
        full_desc = desc_body
        if steps:
            full_desc += f"\n\nActions: {steps}"
        if risks:
            full_desc += f"\n\nRisks: {risks}"
        if vendors_list:
            full_desc += f"\n\nVendors: {vendors_list}"

        savings_str = (
            f"${savings_low:,.0f} – ${savings_high:,.0f}"
            if savings_low and savings_high
            else f"${max(savings_low, savings_high):,.0f}"
        )

        row_num = 4 + i
        ws.append([
            str(opp.get("rank", i + 1)),
            opp.get("title", ""),
            full_desc,
            f"${current:,.0f}" if current else "",
            savings_str,
            opp.get("timeline", ""),
        ])
        ws.row_dimensions[row_num].height = 110

        for col_idx, cell in enumerate(ws[row_num], 1):
            cell.font = BODY_FONT
            cell.fill = opp_fills[i]
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if col_idx == 1:
                cell.font = Font(bold=True, size=16)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            if col_idx == 2:
                cell.font = BOLD_FONT

    # Combined savings row
    total_low  = sum(o.get("savings_low_usd",  0) for o in opportunities[:3])
    total_high = sum(o.get("savings_high_usd", 0) for o in opportunities[:3])
    ws.append([""])
    ws.append(["", "COMBINED ESTIMATED ANNUAL SAVINGS", "", "",
               f"${total_low:,.0f} – ${total_high:,.0f}", ""])
    r = ws.max_row
    ws[f"B{r}"].font = SECTION_FONT
    ws[f"E{r}"].font = Font(bold=True, size=12, color="375623")
    ws.row_dimensions[r].height = 24


# ── Tab 3: Methodology ────────────────────────────────────────────────────────
def _build_methodology_tab(wb: openpyxl.Workbook, insights: dict, qa_report: dict | None) -> None:
    ws = wb.create_sheet("Methodology")
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 88

    total_spend = insights.get("total_spend_usd", 0)
    total_v     = insights.get("total_vendors", 0)
    analysis_dt = insights.get("analysis_date", "")
    rec_summary = insights.get("recommendation_summary", {})

    qa_ok      = qa_report.get("ok", "N/A")      if qa_report else "N/A"
    qa_warn    = qa_report.get("warn", "N/A")    if qa_report else "N/A"
    qa_err     = qa_report.get("error", "N/A")   if qa_report else "N/A"
    qa_reclass = qa_report.get("reclassified", "N/A") if qa_report else "N/A"

    sections = [
        ("VENDOR SPEND ANALYSIS — METHODOLOGY", None, "title"),
        ("", None, "spacer"),
        ("Overview", None, "section"),
        ("", (
            f"This analysis covers {total_v} vendors representing "
            f"${total_spend:,.0f} in trailing twelve-month accounts payable spend. "
            "Vendors were sourced from the provided input file. The objective was to classify "
            "each vendor by department, provide a factual one-line description, and assign a "
            "cost-optimization recommendation to support executive decision-making."
        ), "body"),
        ("", None, "spacer"),
        ("Data Source", None, "section"),
        ("Analysis date",    analysis_dt, "row"),
        ("Total vendors",    str(total_v), "row"),
        ("Total TTM spend",  f"${total_spend:,.0f}", "row"),
        ("Terminate",        str(rec_summary.get("Terminate", 0)), "row"),
        ("Consolidate",      str(rec_summary.get("Consolidate", 0)), "row"),
        ("Optimize",         str(rec_summary.get("Optimize", 0)), "row"),
        ("", None, "spacer"),
        ("5-Step Pipeline", None, "section"),
        ("1 — Fetch", (
            "Loads vendor data from a CSV file or Google Sheets URL/ID via public export. "
            "Auto-detects name and spend columns. Saves a timestamped copy as vendors_raw.csv."
        ), "row"),
        ("2 — Research", (
            "Every vendor submitted to Claude (claude-sonnet-4-6) in batches of 50 for a "
            "training-knowledge pass. Claude classifies confidence as HIGH (well-known company) "
            "or LOW (local/obscure). LOW-confidence vendors above the spend threshold receive an "
            "additional DuckDuckGo web lookup for an independent description snippet. "
            "Results cached to vendors_researched.json."
        ), "row"),
        ("3 — Classify", (
            "Vendors classified via Claude API with: (a) department definitions and recommendation "
            "criteria in the system prompt; (b) research context injected per vendor to ground "
            "descriptions in verified facts; (c) prompt caching (cache_control: ephemeral) on the "
            "system prompt — saves ~85% of input tokens after the first batch; "
            "(d) crash recovery: progress saved after each batch."
        ), "row"),
        ("4 — QA Review", (
            "Second Claude pass reviews every classification as a senior procurement auditor. "
            "Criteria: department fit, description quality, recommendation consistency, factual "
            "accuracy. Severity: ok / warn / error. Error-flagged vendors are automatically "
            "re-classified with QA feedback injected as context. All changes logged."
        ), "row"),
        ("5 — Synthesize", (
            "Claude analyzes the full classified dataset and generates the Top 3 opportunities "
            "and executive memo from actual data. No hardcoded outputs — all figures and vendor "
            "references are derived from the classification results."
        ), "row"),
        ("", None, "spacer"),
        ("QA Statistics", None, "section"),
        ("Passed (ok)",            str(qa_ok), "row"),
        ("Warnings (warn)",        str(qa_warn), "row"),
        ("Errors flagged",         str(qa_err), "row"),
        ("Re-classified after QA", str(qa_reclass), "row"),
        ("", None, "spacer"),
        ("Recommendation Framework", None, "section"),
        ("Terminate", (
            "No identifiable recurring business value; one-off purchases; spend under $500 with "
            "no recurring purpose; or clearly superseded by a higher-spend entry."
        ), "row"),
        ("Consolidate", (
            "Two or more vendors provide overlapping/identical services. Both flagged; "
            "recommendation_note names the specific duplicate."
        ), "row"),
        ("Optimize", (
            "Strategic and necessary vendor — spend is high relative to market benchmarks, "
            "contract should be renegotiated, or usage right-sized."
        ), "row"),
        ("", None, "spacer"),
        ("Tools", None, "section"),
        ("Language",      "Python 3.9+", "row"),
        ("AI model",      "claude-sonnet-4-6 (Anthropic)", "row"),
        ("SDK",           "anthropic (latest)", "row"),
        ("Output",        "openpyxl", "row"),
        ("Web lookup",    "DuckDuckGo Instant Answer API", "row"),
        ("CLI",           "Claude Code (Anthropic)", "row"),
        ("", None, "spacer"),
        ("Limitations", None, "section"),
        ("", (
            "1. No contract terms, renewal dates, or seat/usage data available — savings "
            "estimates are based on industry benchmarks.\n"
            "2. Spend figures represent AP payments; may differ from contracted amounts due "
            "to timing, credits, or prepayments.\n"
            "3. Analysis does not cover vendor performance, SLA compliance, or relationship risk.\n"
            "4. Classification confidence is noted per vendor; LOW-confidence items should "
            "receive human review before acting on recommendations."
        ), "body"),
    ]

    for label, value, style in sections:
        ws.append([label] if value is None else [label, value])
        r = ws.max_row
        if style == "title":
            ws[f"A{r}"].font = TITLE_FONT
            ws.row_dimensions[r].height = 30
        elif style == "section":
            ws[f"A{r}"].font = SECTION_FONT
            ws.row_dimensions[r].height = 22
        elif style == "spacer":
            ws.row_dimensions[r].height = 8
        elif style == "row":
            ws[f"A{r}"].font = BOLD_FONT
            ws[f"A{r}"].alignment = Alignment(vertical="top", wrap_text=True)
            ws[f"B{r}"].font = BODY_FONT
            ws[f"B{r}"].alignment = Alignment(vertical="top", wrap_text=True)
            ws.row_dimensions[r].height = 60
        elif style == "body":
            ws[f"B{r}"].font = BODY_FONT
            ws[f"B{r}"].alignment = Alignment(vertical="top", wrap_text=True)
            ws.row_dimensions[r].height = 80


def _actions_str(value) -> str:
    """Normalise immediate_actions to a plain multi-line string for Excel."""
    if isinstance(value, list):
        return "\n".join(f"• {s}" for s in value if str(s).strip())
    return str(value) if value else ""


def _steps_str(value) -> str:
    """Normalise implementation_steps (list or string) to a plain string."""
    if isinstance(value, list):
        return "\n".join(f"• {s}" for s in value if str(s).strip())
    return str(value) if value else ""


# ── Tab 4: Executive Memo ──────────────────────────────────────────────────────
def _build_executive_memo_tab(wb: openpyxl.Workbook, insights: dict) -> None:
    ws = wb.create_sheet("Recommendations")
    ws.column_dimensions["A"].width = 110

    memo = insights.get("executive_memo", {})
    opps = insights.get("opportunities", [])

    total_low  = sum(o.get("savings_low_usd",  0) for o in opps[:3])
    total_high = sum(o.get("savings_high_usd", 0) for o in opps[:3])

    lines = [
        ("MEMORANDUM", "title"),
        ("", "spacer"),
        (f"TO:\t\t{memo.get('to', 'Chief Executive Officer, Chief Financial Officer')}", "header"),
        (f"FROM:\t\t{memo.get('from', 'VP of Operations')}", "header"),
        (f"DATE:\t\t{memo.get('date', insights.get('analysis_date', ''))}", "header"),
        (f"RE:\t\t{memo.get('subject', 'Vendor Spend Analysis — Strategic Cost Reduction')}", "header"),
        ("", "spacer"),
        ("─" * 100, "divider"),
        ("", "spacer"),
        ("EXECUTIVE SUMMARY", "section"),
        (memo.get("executive_summary", ""), "body"),
        ("", "spacer"),
        ("STRATEGIC OPPORTUNITIES", "section"),
    ]

    for opp in opps[:3]:
        savings_low  = opp.get("savings_low_usd",  0)
        savings_high = opp.get("savings_high_usd", 0)
        savings_str  = f"${savings_low:,.0f} – ${savings_high:,.0f}"
        lines.append(("", "spacer"))
        lines.append((f"{opp.get('rank','')}.  {opp.get('title','')} — Estimated Savings: {savings_str}/year", "bold"))
        body = opp.get("description", "")
        steps = _steps_str(opp.get("implementation_steps", ""))
        if steps:
            body += f"\n\nActions:\n{steps}"
        if opp.get("risks"):
            body += f"\n\nRisk: {opp['risks']}"
        lines.append((body, "body"))

    lines += [
        ("", "spacer"),
        ("─" * 100, "divider"),
        ("", "spacer"),
        ("", "spacer"),
        ("RECOMMENDED IMMEDIATE ACTIONS (30-Day Sprint)", "section"),
        (_actions_str(memo.get("immediate_actions", "")), "body"),
        ("", "spacer"),
        ("─" * 100, "divider"),
        ("", "spacer"),
        (f"TOTAL ESTIMATED ANNUAL SAVINGS:  ${total_low:,.0f} – ${total_high:,.0f}", "bold"),
    ]

    for text, style in lines:
        ws.append([text])
        r = ws.max_row
        if style == "title":
            ws[f"A{r}"].font = TITLE_FONT
            ws.row_dimensions[r].height = 32
        elif style == "section":
            ws[f"A{r}"].font = SECTION_FONT
            ws.row_dimensions[r].height = 22
        elif style == "header":
            ws[f"A{r}"].font = MEMO_HDR_FONT
            ws.row_dimensions[r].height = 18
        elif style == "bold":
            ws[f"A{r}"].font = Font(bold=True, size=11)
            ws.row_dimensions[r].height = 18
        elif style == "body":
            ws[f"A{r}"].font = BODY_FONT
            ws[f"A{r}"].alignment = Alignment(wrap_text=True, vertical="top")
            line_count = max(1, text.count("\n") + 1)
            ws.row_dimensions[r].height = min(20 + line_count * 28, 200)
        elif style == "divider":
            ws[f"A{r}"].font = Font(color="CCCCCC", size=9)
            ws.row_dimensions[r].height = 14
        elif style == "spacer":
            ws.row_dimensions[r].height = 8


# ── Public API ────────────────────────────────────────────────────────────────
def build_xlsx(
    vendors: list[dict],
    classifications: list[dict],
    insights: dict,
    qa_report: dict | None = None,
) -> None:
    wb = openpyxl.Workbook()
    _build_vendor_tab(wb, vendors, classifications)
    _build_opportunities_tab(wb, insights)
    _build_methodology_tab(wb, insights, qa_report)
    _build_executive_memo_tab(wb, insights)
    wb.save(OUTPUT_XLSX)
    print(f"  Saved: {OUTPUT_XLSX}")
